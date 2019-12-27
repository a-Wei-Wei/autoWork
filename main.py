import pandas as pd
import openpyxl as ol
import sys
import os
'''
read_excel return dataframe
df.index 是行的集合
df.columns 是列的集合
所需要的索引： 1 3 4 12
'''
file_name = "重庆-事项接口清单一期V0.0.10.xlsx"
file_out_name = "a.xlsx"
file_out_sheet_name = "模板"


def read_rows():
    workbook = ol.load_workbook(file_name)  # 读取excel
    worksheet = workbook.get_sheet_by_name("接口索引")  # 读取Sheet
    rows, cols = worksheet.max_row, worksheet.max_column
    result_list = []
    for i in range(1, rows):
        ce = worksheet.cell(row=i, column=2)
        fill = ce.fill
        if fill.start_color.rgb == "FF92D050":
            result_list.append(i)
    return result_list


def read_rows_value():
    list_result = read_rows()
    df = pd.read_excel(file_name, sheet_name="接口索引")
    list_result_data = []
    for i in list_result:
        list_result_data.append(df.iloc[i-2])
    return list_result_data


def write_excel(module, uri, group_name):
    df = pd.read_excel(file_out_name, sheet_name=file_out_sheet_name)
    columns = df.columns
    list_result_data = read_rows_value()
    if len(list_result_data) == 0:
        print("无绿色标注部分。。。")
    else:
        for item in list_result_data:
            new_row = {}
            for i in columns:
                new_row.setdefault(i, "")
            keys = list(new_row.keys())
            new_row[keys[1]] = "GA"
            new_row[keys[2]] = module
            new_row[keys[3]] = item[1]
            new_row[keys[4]] = item[uri]
            new_row[keys[5]] = item[12]
            new_row[keys[6]] = "500000000000"
            new_row[keys[7]] = "否"
            new_row[keys[8]] = "否"
            new_row[keys[9]] = group_name
            new_row[keys[12]] = "H5->GSP->P5"
            new_row[keys[13]] = "否"
            new_row[keys[14]] = "2020/01"
            df = df.append(new_row, ignore_index=True)
        # print(new_row)
        df.to_excel(file_out_name, sheet_name=file_out_sheet_name, index=False)


if __name__ == '__main__':
    # fuc()
    # print(o)
    if len(sys.argv) <= 1:
        print("请输入文件名")
    else:
        file_name = sys.argv[1]
        try:
            write_excel("P5OUT", 3, "分行适配组")
            write_excel("GA", 4, "政务后台")
            print("已完成！！！")
        except Exception as ex:
            print(ex)
    os.system("pause")
    # list_re = read_rows()
    # print(len(list_re))
    # read_rows_value()

